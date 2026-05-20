#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合同管理系统 - v5.0
功能：合同管理（29字段）、发票管理、搜索排序、到期预警、应收账款催款
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3
import os
import sys
from datetime import datetime, timedelta
import csv
import hashlib
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 数据库路径
if sys.platform == 'win32':
    DATA_DIR = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'ContractManager')
else:
    DATA_DIR = os.path.join(os.path.expanduser('~'), '.contract_manager')

os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'contracts.db')


class DatabaseManager:
    """数据库管理"""
    
    def __init__(self):
        self.init_database()
    
    def get_connection(self):
        return sqlite3.connect(DB_PATH)
    
    def init_database(self):
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # 合同表 - 29个字段
        cursor.execute("PRAGMA table_info(contracts)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if not columns:
            cursor.execute('''
                CREATE TABLE contracts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    序号 INTEGER,
                    下单日期 TEXT,
                    合同编号 TEXT UNIQUE NOT NULL,
                    项目代码 TEXT,
                    是否变更 TEXT,
                    合同评审日期 TEXT,
                    合同签字日期 TEXT,
                    crm日期 TEXT,
                    合同名称 TEXT,
                    对方单位名称 TEXT,
                    区域 TEXT,
                    销售负责人 TEXT,
                    参考金额 REAL,
                    合同额 REAL,
                    联系人 TEXT,
                    联系电话 TEXT,
                    合同内容 TEXT,
                    到款情况 TEXT,
                    合同起始日期 TEXT,
                    合同终止日期 TEXT,
                    开票日期 TEXT,
                    开票金额 REAL,
                    开票余额 REAL,
                    到款金额 REAL,
                    合同余额 REAL,
                    应收账款 REAL,
                    备注 TEXT,
                    项目预算 REAL,
                    设备数量 INTEGER,
                    催款状态 TEXT DEFAULT '未催款',
                    催款日期 TEXT,
                    催款备注 TEXT,
                    数据哈希 TEXT,
                    创建时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        else:
            # 升级数据库 - 添加缺失字段
            required_columns = {
                '序号': 'INTEGER', '下单日期': 'TEXT', '是否变更': 'TEXT',
                '合同评审日期': 'TEXT', '合同签字日期': 'TEXT', 'crm日期': 'TEXT',
                '销售负责人': 'TEXT', '参考金额': 'REAL', '合同额': 'REAL',
                '联系人': 'TEXT', '联系电话': 'TEXT', '到款情况': 'TEXT',
                '开票日期': 'TEXT', '开票金额': 'REAL', '开票余额': 'REAL',
                '到款金额': 'REAL', '合同余额': 'REAL', '应收账款': 'REAL',
                '备注': 'TEXT', '项目预算': 'REAL', '设备数量': 'INTEGER',
                '数据哈希': 'TEXT'
            }
            
            for col_name, col_type in required_columns.items():
                if col_name not in columns:
                    try:
                        cursor.execute(f'ALTER TABLE contracts ADD COLUMN {col_name} {col_type}')
                    except:
                        pass
            
            # 如果有旧的"销售人员"字段，迁移到"销售负责人"
            if '销售人员' in columns and '销售负责人' not in columns:
                try:
                    cursor.execute('ALTER TABLE contracts ADD COLUMN 销售负责人 TEXT')
                    cursor.execute('UPDATE contracts SET 销售负责人 = 销售人员 WHERE 销售负责人 IS NULL')
                except:
                    pass
        
        # 发票表
        cursor.execute("PRAGMA table_info(invoices_new)")
        if not cursor.fetchall():
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS invoices_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    开票日期 TEXT,
                    合同号 TEXT,
                    付款单位名称 TEXT,
                    代码 TEXT,
                    发票金额 REAL,
                    发票项目 TEXT,
                    类型 TEXT,
                    发票类型 TEXT,
                    除税 REAL,
                    备注 TEXT,
                    数据哈希 TEXT,
                    创建时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        
        # 催款记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS collection_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                合同编号 TEXT NOT NULL,
                催款日期 TEXT,
                催款方式 TEXT,
                联系人 TEXT,
                催款内容 TEXT,
                对方反馈 TEXT,
                催款结果 TEXT,
                FOREIGN KEY (合同编号) REFERENCES contracts(合同编号)
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def generate_hash(self, data, fields):
        """生成数据哈希用于去重"""
        hash_str = '|'.join([str(data.get(f, '')) for f in fields])
        return hashlib.md5(hash_str.encode('utf-8')).hexdigest()
    
    def check_duplicate_contract(self, data):
        """检查合同是否重复（所有字段完全一致）"""
        fields = ['序号', '下单日期', '合同编号', '项目代码', '是否变更', '合同评审日期',
                 '合同签字日期', 'crm日期', '合同名称', '对方单位名称', '区域', '销售负责人',
                 '参考金额', '合同额', '联系人', '联系电话', '合同内容', '到款情况',
                 '合同起始日期', '合同终止日期', '开票日期', '开票金额', '开票余额',
                 '到款金额', '合同余额', '应收账款', '备注', '项目预算', '设备数量']
        
        data_hash = self.generate_hash(data, fields)
        
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM contracts WHERE 数据哈希 = ?', (data_hash,))
        result = cursor.fetchone()
        conn.close()
        
        return result is not None, data_hash
    
    def add_contract(self, data):
        """添加合同"""
        is_dup, data_hash = self.check_duplicate_contract(data)
        if is_dup:
            return False, "数据完全重复，已跳过"
        
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO contracts 
                (序号, 下单日期, 合同编号, 项目代码, 是否变更, 合同评审日期,
                 合同签字日期, crm日期, 合同名称, 对方单位名称, 区域, 销售负责人,
                 参考金额, 合同额, 联系人, 联系电话, 合同内容, 到款情况,
                 合同起始日期, 合同终止日期, 开票日期, 开票金额, 开票余额,
                 到款金额, 合同余额, 应收账款, 备注, 项目预算, 设备数量, 数据哈希)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('序号'), data.get('下单日期'), data.get('合同编号'),
                data.get('项目代码'), data.get('是否变更'), data.get('合同评审日期'),
                data.get('合同签字日期'), data.get('crm日期'), data.get('合同名称'),
                data.get('对方单位名称'), data.get('区域'), data.get('销售负责人'),
                data.get('参考金额'), data.get('合同额'), data.get('联系人'),
                data.get('联系电话'), data.get('合同内容'), data.get('到款情况'),
                data.get('合同起始日期'), data.get('合同终止日期'), data.get('开票日期'),
                data.get('开票金额'), data.get('开票余额'), data.get('到款金额'),
                data.get('合同余额'), data.get('应收账款'), data.get('备注'),
                data.get('项目预算'), data.get('设备数量'), data_hash
            ))
            conn.commit()
            return True, "添加成功"
        except sqlite3.IntegrityError as e:
            return False, f"合同编号已存在: {data.get('合同编号')}"
        except Exception as e:
            return False, f"添加失败: {str(e)}"
        finally:
            conn.close()
    
    def update_contract(self, contract_no, data):
        """更新合同"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE contracts SET 
                序号=?, 下单日期=?, 项目代码=?, 是否变更=?, 合同评审日期=?,
                合同签字日期=?, crm日期=?, 合同名称=?, 对方单位名称=?, 区域=?,
                销售负责人=?, 参考金额=?, 合同额=?, 联系人=?, 联系电话=?,
                合同内容=?, 到款情况=?, 合同起始日期=?, 合同终止日期=?,
                开票日期=?, 开票金额=?, 开票余额=?, 到款金额=?, 合同余额=?,
                应收账款=?, 备注=?, 项目预算=?, 设备数量=?
                WHERE 合同编号=?
            ''', (
                data.get('序号'), data.get('下单日期'), data.get('项目代码'),
                data.get('是否变更'), data.get('合同评审日期'), data.get('合同签字日期'),
                data.get('crm日期'), data.get('合同名称'), data.get('对方单位名称'),
                data.get('区域'), data.get('销售负责人'), data.get('参考金额'),
                data.get('合同额'), data.get('联系人'), data.get('联系电话'),
                data.get('合同内容'), data.get('到款情况'), data.get('合同起始日期'),
                data.get('合同终止日期'), data.get('开票日期'), data.get('开票金额'),
                data.get('开票余额'), data.get('到款金额'), data.get('合同余额'),
                data.get('应收账款'), data.get('备注'), data.get('项目预算'),
                data.get('设备数量'), contract_no
            ))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()
    
    def get_contract_by_no(self, contract_no):
        """根据合同编号获取合同"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM contracts WHERE 合同编号=?', (contract_no,))
        row = cursor.fetchone()
        conn.close()
        return row
    
    def delete_contract(self, contract_no):
        """删除合同"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM collection_records WHERE 合同编号=?', (contract_no,))
            cursor.execute('DELETE FROM contracts WHERE 合同编号=?', (contract_no,))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()
    
    def get_contracts(self, year=None, region=None, salesperson=None, search_text=None):
        """获取合同列表"""
        conn = self.get_connection()
        
        query = 'SELECT * FROM contracts'
        conditions = []
        params = []
        
        if year:
            conditions.append("strftime('%Y', 合同签字日期) = ?")
            params.append(str(year))
        if region:
            conditions.append("区域 = ?")
            params.append(region)
        if salesperson:
            conditions.append("销售负责人 = ?")
            params.append(salesperson)
        if search_text:
            conditions.append('''(
                合同编号 LIKE ? OR 
                项目代码 LIKE ? OR 
                合同名称 LIKE ? OR 
                对方单位名称 LIKE ? OR
                销售负责人 LIKE ? OR
                联系人 LIKE ?
            )''')
            search_pattern = f'%{search_text}%'
            params.extend([search_pattern] * 6)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY 创建时间 DESC"
        
        cursor = conn.cursor()
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_expiring_contracts(self, days=30):
        """获取即将到期的合同"""
        conn = self.get_connection()
        today = datetime.now().strftime('%Y-%m-%d')
        future = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM contracts 
            WHERE 合同终止日期 >= ? AND 合同终止日期 <= ?
            ORDER BY 合同终止日期 ASC
        ''', (today, future))
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_expired_contracts(self):
        """获取已到期的合同"""
        conn = self.get_connection()
        today = datetime.now().strftime('%Y-%m-%d')
        
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM contracts 
            WHERE 合同终止日期 < ?
            ORDER BY 合同终止日期 DESC
        ''', (today,))
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_receivables_by_year(self, year):
        """获取指定年份的应收账款"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT * FROM contracts 
            WHERE strftime('%Y', 合同签字日期) = ? AND 应收账款 > 0
            ORDER BY 合同签字日期 DESC
        ''', (str(year),))
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_years_with_receivables(self):
        """获取有应收账款的年份列表"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT strftime('%Y', 合同签字日期) as 年度
            FROM contracts 
            WHERE 合同签字日期 IS NOT NULL AND 合同签字日期 != '' AND 应收账款 > 0
            ORDER BY 年度 DESC
        ''')
        rows = cursor.fetchall()
        conn.close()
        return [row[0] for row in rows if row[0]]
    
    # 发票管理
    def check_duplicate_invoice(self, data):
        """检查发票是否重复"""
        fields = ['开票日期', '合同号', '付款单位名称', '代码', '发票金额',
                 '发票项目', '类型', '发票类型', '除税', '备注']
        
        data_hash = self.generate_hash(data, fields)
        
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM invoices_new WHERE 数据哈希 = ?', (data_hash,))
        result = cursor.fetchone()
        conn.close()
        
        return result is not None, data_hash
    
    def add_invoice_new(self, data):
        """添加发票"""
        is_dup, data_hash = self.check_duplicate_invoice(data)
        if is_dup:
            return False, "数据完全重复，已跳过"
        
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO invoices_new 
                (开票日期, 合同号, 付款单位名称, 代码, 发票金额, 发票项目,
                 类型, 发票类型, 除税, 备注, 数据哈希)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('开票日期'), data.get('合同号'), data.get('付款单位名称'),
                data.get('代码'), data.get('发票金额'), data.get('发票项目'),
                data.get('类型'), data.get('发票类型'), data.get('除税'),
                data.get('备注'), data_hash
            ))
            conn.commit()
            return True, "添加成功"
        except Exception as e:
            return False, f"添加失败: {str(e)}"
        finally:
            conn.close()
    
    def update_invoice_new(self, invoice_id, data):
        """更新发票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE invoices_new SET 
                开票日期=?, 合同号=?, 付款单位名称=?, 代码=?, 发票金额=?,
                发票项目=?, 类型=?, 发票类型=?, 除税=?, 备注=?
                WHERE id=?
            ''', (
                data.get('开票日期'), data.get('合同号'), data.get('付款单位名称'),
                data.get('代码'), data.get('发票金额'), data.get('发票项目'),
                data.get('类型'), data.get('发票类型'), data.get('除税'),
                data.get('备注'), invoice_id
            ))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()
    
    def get_invoice_by_id(self, invoice_id):
        """根据ID获取发票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM invoices_new WHERE id=?', (invoice_id,))
        row = cursor.fetchone()
        conn.close()
        return row
    
    def delete_invoice_new(self, invoice_id):
        """删除发票"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM invoices_new WHERE id=?', (invoice_id,))
        conn.commit()
        conn.close()
        return cursor.rowcount > 0
    
    def get_invoices_new(self, search_text=None):
        """获取发票列表"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        if search_text:
            cursor.execute('''
                SELECT * FROM invoices_new 
                WHERE 开票日期 LIKE ? OR 合同号 LIKE ? OR 付款单位名称 LIKE ? OR 代码 LIKE ?
                ORDER BY 创建时间 DESC
            ''', (f'%{search_text}%', f'%{search_text}%', f'%{search_text}%', f'%{search_text}%'))
        else:
            cursor.execute('SELECT * FROM invoices_new ORDER BY 创建时间 DESC')
        
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def update_collection_status(self, contract_no, status, date='', note=''):
        """更新催款状态"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE contracts SET 催款状态=?, 催款日期=?, 催款备注=?
            WHERE 合同编号=?
        ''', (status, date, note, contract_no))
        conn.commit()
        conn.close()
    
    def add_collection_record(self, contract_no, date, method, contact, content, feedback, result):
        """添加催款记录"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO collection_records 
            (合同编号, 催款日期, 催款方式, 联系人, 催款内容, 对方反馈, 催款结果)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (contract_no, date, method, contact, content, feedback, result))
        conn.commit()
        conn.close()
    
    def get_collection_records(self, contract_no):
        """获取催款记录"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 催款日期, 催款方式, 联系人, 催款内容, 对方反馈, 催款结果
            FROM collection_records
            WHERE 合同编号 = ?
            ORDER BY 催款日期 DESC
        ''', (contract_no,))
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_yearly_stats(self):
        """年度统计"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 
                strftime('%Y', 合同签字日期) as 年度,
                COUNT(*) as 合同数,
                SUM(合同额) as 合同总额,
                SUM(开票金额) as 开票总额,
                SUM(到款金额) as 回款总额
            FROM contracts
            WHERE 合同签字日期 IS NOT NULL AND 合同签字日期 != ''
            GROUP BY strftime('%Y', 合同签字日期)
            ORDER BY 年度 DESC
        ''')
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_region_stats(self, year=None):
        """区域统计"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        if year:
            cursor.execute('''
                SELECT 区域, COUNT(*) as 合同数, SUM(合同额) as 合同总额,
                       SUM(开票金额) as 开票总额, SUM(到款金额) as 回款总额
                FROM contracts
                WHERE strftime('%Y', 合同签字日期) = ?
                GROUP BY 区域
                ORDER BY 合同总额 DESC
            ''', (str(year),))
        else:
            cursor.execute('''
                SELECT 区域, COUNT(*) as 合同数, SUM(合同额) as 合同总额,
                       SUM(开票金额) as 开票总额, SUM(到款金额) as 回款总额
                FROM contracts
                GROUP BY 区域
                ORDER BY 合同总额 DESC
            ''')
        
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def get_salesperson_stats(self, year=None):
        """销售人员统计"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        if year:
            cursor.execute('''
                SELECT 销售负责人, COUNT(*) as 合同数, SUM(合同额) as 合同总额,
                       SUM(开票金额) as 开票总额, SUM(到款金额) as 回款总额
                FROM contracts
                WHERE strftime('%Y', 合同签字日期) = ?
                GROUP BY 销售负责人
                ORDER BY 合同总额 DESC
            ''', (str(year),))
        else:
            cursor.execute('''
                SELECT 销售负责人, COUNT(*) as 合同数, SUM(合同额) as 合同总额,
                       SUM(开票金额) as 开票总额, SUM(到款金额) as 回款总额
                FROM contracts
                GROUP BY 销售负责人
                ORDER BY 合同总额 DESC
            ''')
        
        rows = cursor.fetchall()
        conn.close()
        return rows
    
    def import_contracts_from_file(self, file_path):
        """从文件导入合同"""
        count = 0
        dup_count = 0
        error_count = 0
        
        try:
            if file_path.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # 读取表头
                    headers = []
                    for cell in ws[1]:
                        header = str(cell.value).strip() if cell.value else ''
                        headers.append(header)
                    
                    # 读取数据行
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] and not row[2]:  # 序号和合同编号都为空则跳过
                            continue
                        
                        row_data = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                row_data[header] = row[i]
                        
                        # 字段映射
                        data = {
                            '序号': row_data.get('序号'),
                            '下单日期': self._format_date(row_data.get('下单日期')),
                            '合同编号': str(row_data.get('合同编号', '') or ''),
                            '项目代码': str(row_data.get('项目代码', '') or ''),
                            '是否变更': str(row_data.get('是否变更', '') or ''),
                            '合同评审日期': self._format_date(row_data.get('合同评审日期')),
                            '合同签字日期': self._format_date(row_data.get('合同签字日期') or row_data.get('实际签约日期')),
                            'crm日期': self._format_date(row_data.get('crm日期')),
                            '合同名称': str(row_data.get('合同名称', '') or ''),
                            '对方单位名称': str(row_data.get('对方单位名称', '') or ''),
                            '区域': str(row_data.get('区域', '') or ''),
                            '销售负责人': str(row_data.get('销售负责人', '') or row_data.get('销售人员', '') or ''),
                            '参考金额': self._parse_float(row_data.get('参考金额')),
                            '合同额': self._parse_float(row_data.get('合同额') or row_data.get('合同金额')),
                            '联系人': str(row_data.get('联系人', '') or ''),
                            '联系电话': str(row_data.get('联系电话', '') or ''),
                            '合同内容': str(row_data.get('合同内容', '') or ''),
                            '到款情况': str(row_data.get('到款情况', '') or ''),
                            '合同起始日期': self._format_date(row_data.get('合同起始日期')),
                            '合同终止日期': self._format_date(row_data.get('合同终止日期')),
                            '开票日期': self._format_date(row_data.get('开票日期')),
                            '开票金额': self._parse_float(row_data.get('开票金额')),
                            '开票余额': self._parse_float(row_data.get('开票余额')),
                            '到款金额': self._parse_float(row_data.get('到款金额')),
                            '合同余额': self._parse_float(row_data.get('合同余额')),
                            '应收账款': self._parse_float(row_data.get('应收账款')),
                            '备注': str(row_data.get('备注', '') or ''),
                            '项目预算': self._parse_float(row_data.get('项目预算')),
                            '设备数量': self._parse_int(row_data.get('设备数量'))
                        }
                        
                        if data['合同编号'] and data['合同编号'] != 'None':
                            success, msg = self.add_contract(data)
                            if success:
                                count += 1
                            elif '重复' in msg:
                                dup_count += 1
                            else:
                                error_count += 1
                
                wb.close()
            else:
                # CSV文件
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        data = {
                            '序号': row.get('序号'),
                            '下单日期': row.get('下单日期', ''),
                            '合同编号': row.get('合同编号', ''),
                            '项目代码': row.get('项目代码', ''),
                            '是否变更': row.get('是否变更', ''),
                            '合同评审日期': row.get('合同评审日期', ''),
                            '合同签字日期': row.get('合同签字日期', ''),
                            'crm日期': row.get('crm日期', ''),
                            '合同名称': row.get('合同名称', ''),
                            '对方单位名称': row.get('对方单位名称', ''),
                            '区域': row.get('区域', ''),
                            '销售负责人': row.get('销售负责人', '') or row.get('销售人员', ''),
                            '参考金额': self._parse_float(row.get('参考金额')),
                            '合同额': self._parse_float(row.get('合同额') or row.get('合同金额')),
                            '联系人': row.get('联系人', ''),
                            '联系电话': row.get('联系电话', ''),
                            '合同内容': row.get('合同内容', ''),
                            '到款情况': row.get('到款情况', ''),
                            '合同起始日期': row.get('合同起始日期', ''),
                            '合同终止日期': row.get('合同终止日期', ''),
                            '开票日期': row.get('开票日期', ''),
                            '开票金额': self._parse_float(row.get('开票金额')),
                            '开票余额': self._parse_float(row.get('开票余额')),
                            '到款金额': self._parse_float(row.get('到款金额')),
                            '合同余额': self._parse_float(row.get('合同余额')),
                            '应收账款': self._parse_float(row.get('应收账款')),
                            '备注': row.get('备注', ''),
                            '项目预算': self._parse_float(row.get('项目预算')),
                            '设备数量': self._parse_int(row.get('设备数量'))
                        }
                        
                        if data['合同编号']:
                            success, msg = self.add_contract(data)
                            if success:
                                count += 1
                            elif '重复' in msg:
                                dup_count += 1
                            else:
                                error_count += 1
            
            return count, dup_count, error_count
        except Exception as e:
            raise e
    
    def import_invoices_from_file(self, file_path):
        """从文件导入发票"""
        count = 0
        dup_count = 0
        error_count = 0
        
        try:
            if file_path.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    headers = []
                    for cell in ws[1]:
                        header = str(cell.value).strip() if cell.value else ''
                        headers.append(header)
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row):  # 空行跳过
                            continue
                        
                        row_data = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                row_data[header] = row[i]
                        
                        data = {
                            '开票日期': self._format_date(row_data.get('开票日期')),
                            '合同号': str(row_data.get('合同号', '') or ''),
                            '付款单位名称': str(row_data.get('付款单位名称', '') or ''),
                            '代码': str(row_data.get('代码', '') or ''),
                            '发票金额': self._parse_float(row_data.get('发票金额')),
                            '发票项目': str(row_data.get('发票项目', '') or ''),
                            '类型': str(row_data.get('类型', '') or ''),
                            '发票类型': str(row_data.get('发票类型', '') or ''),
                            '除税': self._parse_float(row_data.get('除税')),
                            '备注': str(row_data.get('备注', '') or '')
                        }
                        
                        success, msg = self.add_invoice_new(data)
                        if success:
                            count += 1
                        elif '重复' in msg:
                            dup_count += 1
                        else:
                            error_count += 1
                
                wb.close()
            
            return count, dup_count, error_count
        except Exception as e:
            raise e
    
    def _format_date(self, date_value):
        """格式化日期"""
        if not date_value:
            return ''
        
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d')
        
        date_str = str(date_value)
        if date_str == 'None' or not date_str.strip():
            return ''
        
        try:
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                try:
                    parsed = datetime.strptime(date_str.strip(), fmt)
                    return parsed.strftime('%Y-%m-%d')
                except:
                    continue
            return date_str
        except:
            return date_str
    
    def _parse_float(self, value):
        """解析浮点数"""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            return float(str(value).replace(',', '').replace(' ', ''))
        except:
            return None
    
    def _parse_int(self, value):
        """解析整数"""
        if value is None:
            return None
        try:
            if isinstance(value, int):
                return value
            return int(str(value).replace(',', '').replace(' ', ''))
        except:
            return None


class ContractManagerApp:
    """主应用"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("合同管理系统 v5.0")
        self.root.geometry("1600x900")
        
        self.db = DatabaseManager()
        
        self.regions = ['北方区', '西北区', '华东区', '华南区', '国际部', '其他']
        self.collection_status = ['未催款', '催款中', '已承诺付款', '已回款', '坏账']
        self.collection_methods = ['电话', '邮件', '上门', '微信', '其他']
        
        self.sort_column = None
        self.sort_reverse = False
        
        self.create_widgets()
        self.load_data()
        self.check_expiring_contracts()
    
    def create_widgets(self):
        # 创建菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 数据菜单
        data_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="数据", menu=data_menu)
        data_menu.add_command(label="备份数据库", command=self.backup_database)
        data_menu.add_command(label="恢复数据库", command=self.restore_database)
        data_menu.add_separator()
        data_menu.add_command(label="导出合同Excel", command=self.export_contracts)
        data_menu.add_command(label="导出发票Excel", command=self.export_invoices)
        
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="关于", command=self.show_about)
        
        # 创建主标签页
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # 合同管理
        contract_frame = ttk.Frame(notebook)
        notebook.add(contract_frame, text='合同管理')
        self.create_contract_tab(contract_frame)
        
        # 发票管理
        invoice_frame = ttk.Frame(notebook)
        notebook.add(invoice_frame, text='发票管理')
        self.create_invoice_tab(invoice_frame)
        
        # 到期预警
        warning_frame = ttk.Frame(notebook)
        notebook.add(warning_frame, text='到期预警')
        self.create_warning_tab(warning_frame)
        
        # 应收账款
        receivable_frame = ttk.Frame(notebook)
        notebook.add(receivable_frame, text='应收账款')
        self.create_receivable_tab(receivable_frame)
        
        # 年度统计
        yearly_frame = ttk.Frame(notebook)
        notebook.add(yearly_frame, text='年度统计')
        self.create_yearly_tab(yearly_frame)
        
        # 区域统计
        region_frame = ttk.Frame(notebook)
        notebook.add(region_frame, text='区域统计')
        self.create_region_tab(region_frame)
        
        # 销售人员统计
        sales_frame = ttk.Frame(notebook)
        notebook.add(sales_frame, text='销售人员统计')
        self.create_salesperson_tab(sales_frame)
    
    def create_contract_tab(self, parent):
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(toolbar, text="添加合同", command=self.add_contract).pack(side='left', padx=2)
        ttk.Button(toolbar, text="修改合同", command=self.edit_contract).pack(side='left', padx=2)
        ttk.Button(toolbar, text="删除合同", command=self.delete_contract).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导入数据", command=self.import_contracts).pack(side='left', padx=2)
        ttk.Button(toolbar, text="刷新", command=self.load_data).pack(side='left', padx=2)
        
        # 搜索和筛选栏
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(filter_frame, text="搜索:").pack(side='left', padx=2)
        self.contract_search_entry = ttk.Entry(filter_frame, width=20)
        self.contract_search_entry.pack(side='left', padx=2)
        self.contract_search_entry.bind('<Return>', lambda e: self.load_data())
        ttk.Button(filter_frame, text="搜索", command=self.load_data).pack(side='left', padx=2)
        ttk.Button(filter_frame, text="清除", command=self.clear_contract_search).pack(side='left', padx=2)
        
        ttk.Separator(filter_frame, orient='vertical').pack(side='left', padx=10, fill='y')
        
        ttk.Label(filter_frame, text="年度:").pack(side='left', padx=2)
        self.contract_year_filter = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.contract_year_filter['values'] = ['全部'] + [str(y) for y in range(2030, 2019, -1)]
        self.contract_year_filter.set('全部')
        self.contract_year_filter.pack(side='left', padx=2)
        self.contract_year_filter.bind('<<ComboboxSelected>>', lambda e: self.load_data())
        
        ttk.Label(filter_frame, text="区域:").pack(side='left', padx=2)
        self.contract_region_filter = ttk.Combobox(filter_frame, width=12, state='readonly')
        self.contract_region_filter['values'] = ['全部'] + self.regions
        self.contract_region_filter.set('全部')
        self.contract_region_filter.pack(side='left', padx=2)
        self.contract_region_filter.bind('<<ComboboxSelected>>', lambda e: self.load_data())
        
        ttk.Label(filter_frame, text="销售负责人:").pack(side='left', padx=2)
        self.contract_salesperson_filter = ttk.Combobox(filter_frame, width=12, state='readonly')
        self.contract_salesperson_filter['values'] = ['全部']
        self.contract_salesperson_filter.set('全部')
        self.contract_salesperson_filter.pack(side='left', padx=2)
        self.contract_salesperson_filter.bind('<<ComboboxSelected>>', lambda e: self.load_data())
        
        # 合同列表 - 29个字段
        columns = ('序号', '下单日期', '合同编号', '项目代码', '是否变更', '合同评审日期',
                  '合同签字日期', 'crm日期', '合同名称', '对方单位名称', '区域', '销售负责人',
                  '参考金额', '合同额', '联系人', '联系电话', '合同内容', '到款情况',
                  '合同起始日期', '合同终止日期', '开票日期', '开票金额', '开票余额',
                  '到款金额', '合同余额', '应收账款', '备注', '项目预算', '设备数量')
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.contract_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=25)
        
        # 设置列
        for col in columns:
            self.contract_tree.heading(col, text=col, command=lambda c=col: self.sort_contract_by_column(c))
            self.contract_tree.column(col, width=100, anchor='center')
        
        # 调整部分列宽
        self.contract_tree.column('合同名称', width=150)
        self.contract_tree.column('对方单位名称', width=150)
        self.contract_tree.column('备注', width=150)
        
        # 滚动条
        scrollbar_y = ttk.Scrollbar(tree_frame, orient='vertical', command=self.contract_tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.contract_tree.xview)
        self.contract_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.contract_tree.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
    
    def create_invoice_tab(self, parent):
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(toolbar, text="添加发票", command=self.add_invoice).pack(side='left', padx=2)
        ttk.Button(toolbar, text="修改发票", command=self.edit_invoice).pack(side='left', padx=2)
        ttk.Button(toolbar, text="删除发票", command=self.delete_invoice).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导入数据", command=self.import_invoices).pack(side='left', padx=2)
        ttk.Button(toolbar, text="刷新", command=self.load_invoices).pack(side='left', padx=2)
        
        # 搜索栏
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side='left', padx=2)
        self.invoice_search_entry = ttk.Entry(search_frame, width=30)
        self.invoice_search_entry.pack(side='left', padx=2)
        self.invoice_search_entry.bind('<Return>', lambda e: self.load_invoices())
        ttk.Button(search_frame, text="搜索", command=self.load_invoices).pack(side='left', padx=2)
        ttk.Button(search_frame, text="清除", command=self.clear_invoice_search).pack(side='left', padx=2)
        
        # 发票列表
        columns = ('开票日期', '合同号', '付款单位名称', '代码', '发票金额', '发票项目',
                  '类型', '发票类型', '除税', '备注')
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.invoice_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=30)
        
        for col in columns:
            self.invoice_tree.heading(col, text=col, command=lambda c=col: self.sort_invoice_by_column(c))
            self.invoice_tree.column(col, width=120, anchor='center')
        
        self.invoice_tree.column('付款单位名称', width=180)
        self.invoice_tree.column('备注', width=150)
        
        scrollbar_y = ttk.Scrollbar(tree_frame, orient='vertical', command=self.invoice_tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.invoice_tree.xview)
        self.invoice_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.invoice_tree.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
    
    def create_warning_tab(self, parent):
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(toolbar, text="刷新", command=self.load_warning_data).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导出预警列表", command=self.export_warning_list).pack(side='left', padx=2)
        
        # 快速过滤按钮
        ttk.Separator(toolbar, orient='vertical').pack(side='left', padx=10, fill='y')
        ttk.Label(toolbar, text="快速筛选:").pack(side='left', padx=2)
        
        self.warning_filter_var = tk.StringVar(value='all')
        
        ttk.Radiobutton(toolbar, text="全部", variable=self.warning_filter_var, 
                       value='all', command=self.load_warning_data).pack(side='left', padx=2)
        ttk.Radiobutton(toolbar, text="30天内到期", variable=self.warning_filter_var, 
                       value='yellow', command=self.load_warning_data).pack(side='left', padx=2)
        ttk.Radiobutton(toolbar, text="已到期0-30天", variable=self.warning_filter_var, 
                       value='orange', command=self.load_warning_data).pack(side='left', padx=2)
        ttk.Radiobutton(toolbar, text="超期30天以上", variable=self.warning_filter_var, 
                       value='red', command=self.load_warning_data).pack(side='left', padx=2)
        
        # 搜索栏
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(search_frame, text="搜索:").pack(side='left', padx=2)
        self.warning_search_entry = ttk.Entry(search_frame, width=30)
        self.warning_search_entry.pack(side='left', padx=2)
        self.warning_search_entry.bind('<Return>', lambda e: self.load_warning_data())
        ttk.Button(search_frame, text="搜索", command=self.load_warning_data).pack(side='left', padx=2)
        ttk.Button(search_frame, text="清除", command=self.clear_warning_search).pack(side='left', padx=2)
        
        # 说明
        info_frame = ttk.Frame(parent)
        info_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(info_frame, text="预警说明：", font=('Arial', 10, 'bold')).pack(side='left')
        ttk.Label(info_frame, text="黄色 - 30天内到期  |  橙色 - 已到期0-30天  |  红色 - 超期30天以上", 
                 foreground='gray').pack(side='left', padx=10)
        
        # 预警列表
        columns = ('合同编号', '合同名称', '对方单位', '区域', '销售负责人', '合同额',
                  '终止日期', '剩余天数', '开票金额', '到款金额', '应收账款')
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.warning_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=30)
        
        for col in columns:
            self.warning_tree.heading(col, text=col)
            self.warning_tree.column(col, width=120, anchor='center')
        
        # 配置标签颜色
        self.warning_tree.tag_configure('warning_yellow', background='#FFF3CD')
        self.warning_tree.tag_configure('warning_orange', background='#FFE5B4')
        self.warning_tree.tag_configure('warning_red', background='#FFCCCC')
        
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.warning_tree.yview)
        self.warning_tree.configure(yscrollcommand=scrollbar.set)
        
        self.warning_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
    
    def create_receivable_tab(self, parent):
        # 创建年份Notebook
        self.receivable_notebook = ttk.Notebook(parent)
        self.receivable_notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.load_receivable_tabs()
    
    def load_receivable_tabs(self):
        """加载应收账款的年份标签页"""
        for tab in self.receivable_notebook.tabs():
            self.receivable_notebook.forget(tab)
        
        years = self.db.get_years_with_receivables()
        
        if not years:
            empty_frame = ttk.Frame(self.receivable_notebook)
            self.receivable_notebook.add(empty_frame, text='暂无数据')
            ttk.Label(empty_frame, text="当前没有应收账款数据", font=('Arial', 14)).pack(pady=50)
            return
        
        for year in years:
            year_frame = ttk.Frame(self.receivable_notebook)
            self.receivable_notebook.add(year_frame, text=f'{year}年')
            self.create_receivable_year_tab(year_frame, year)
    
    def create_receivable_year_tab(self, parent, year):
        """创建指定年份的应收账款标签页"""
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(toolbar, text="刷新", command=lambda: self.load_receivable_data(parent, year)).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导出Excel", command=lambda: self.export_receivable_excel(year)).pack(side='left', padx=2)
        
        # 账龄分析
        aging_frame = ttk.LabelFrame(parent, text="账龄分析")
        aging_frame.pack(fill='x', padx=5, pady=5)
        
        self.aging_labels = {}
        aging_periods = ['30天内', '31-60天', '61-90天', '91-180天', '180天以上']
        
        for i, period in enumerate(aging_periods):
            ttk.Label(aging_frame, text=f"{period}:").grid(row=0, column=i*2, padx=10, pady=5)
            self.aging_labels[period] = ttk.Label(aging_frame, text="0.00 元", font=('Arial', 10, 'bold'))
            self.aging_labels[period].grid(row=0, column=i*2+1, padx=5, pady=5)
        
        # 应收账款列表
        columns = ('合同编号', '合同名称', '对方单位', '区域', '销售负责人', '合同额',
                  '签字日期', '账龄(天)', '开票金额', '到款金额', '应收金额', '催款状态', '操作')
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor='center')
        
        tree.column('合同名称', width=150)
        tree.column('对方单位', width=150)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        if not hasattr(self, 'receivable_trees'):
            self.receivable_trees = {}
        self.receivable_trees[year] = tree
        
        self.load_receivable_data(parent, year)
    
    def load_receivable_data(self, parent, year):
        """加载指定年份的应收账款数据"""
        if year not in self.receivable_trees:
            return
        
        tree = self.receivable_trees[year]
        
        for item in tree.get_children():
            tree.delete(item)
        
        rows = self.db.get_receivables_by_year(year)
        
        today = datetime.now()
        aging_data = {'30天内': 0, '31-60天': 0, '61-90天': 0, '91-180天': 0, '180天以上': 0}
        
        for row in rows:
            # 根据字段位置提取数据
            序号 = row[0] if len(row) > 0 else None
            下单日期 = row[1] if len(row) > 1 else None
            合同编号 = row[3] if len(row) > 3 else ''
            合同名称 = row[9] if len(row) > 9 else ''
            对方单位 = row[10] if len(row) > 10 else ''
            区域 = row[11] if len(row) > 11 else ''
            销售负责人 = row[12] if len(row) > 12 else ''
            合同额 = row[14] if len(row) > 14 else 0
            合同签字日期 = row[7] if len(row) > 7 else ''
            开票金额 = row[23] if len(row) > 23 else 0
            到款金额 = row[25] if len(row) > 25 else 0
            应收账款 = row[27] if len(row) > 27 else 0
            催款状态 = row[31] if len(row) > 31 else '未催款'
            
            # 计算账龄
            aging_days = 0
            if 合同签字日期:
                try:
                    sign_dt = datetime.strptime(合同签字日期, '%Y-%m-%d')
                    aging_days = (today - sign_dt).days
                except:
                    pass
            
            # 账龄分组
            if aging_days <= 30:
                aging_data['30天内'] += 应收账款 or 0
            elif aging_days <= 60:
                aging_data['31-60天'] += 应收账款 or 0
            elif aging_days <= 90:
                aging_data['61-90天'] += 应收账款 or 0
            elif aging_days <= 180:
                aging_data['91-180天'] += 应收账款 or 0
            else:
                aging_data['180天以上'] += 应收账款 or 0
            
            tree.insert('', 'end', values=(
                合同编号, 合同名称 or '', 对方单位, 区域, 销售负责人 or '',
                f'{合同额:,.2f}' if 合同额 else '0.00',
                合同签字日期 or '', aging_days,
                f'{开票金额:,.2f}' if 开票金额 else '0.00',
                f'{到款金额:,.2f}' if 到款金额 else '0.00',
                f'{应收账款:,.2f}' if 应收账款 else '0.00',
                催款状态 or '未催款', '操作'
            ))
        
        # 更新账龄分析
        if hasattr(self, 'aging_labels'):
            for period, amount in aging_data.items():
                if period in self.aging_labels:
                    self.aging_labels[period].config(text=f'{amount:,.2f} 元')
    
    def create_yearly_tab(self, parent):
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        columns = ('年度', '合同数', '合同总额', '开票总额', '回款总额', '应收账款')
        self.yearly_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.yearly_tree.heading(col, text=col)
            self.yearly_tree.column(col, width=150, anchor='center')
        
        self.yearly_tree.pack(fill='both', expand=True)
        
        chart_frame = ttk.LabelFrame(parent, text="年度统计图表")
        chart_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.yearly_fig = Figure(figsize=(12, 4), dpi=100)
        self.yearly_canvas = FigureCanvasTkAgg(self.yearly_fig, master=chart_frame)
        self.yearly_canvas.get_tk_widget().pack(fill='both', expand=True)
    
    def create_region_tab(self, parent):
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(filter_frame, text="年度:").pack(side='left', padx=2)
        self.region_year_filter = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.region_year_filter['values'] = ['全部'] + [str(y) for y in range(2030, 2019, -1)]
        self.region_year_filter.set('全部')
        self.region_year_filter.pack(side='left', padx=2)
        self.region_year_filter.bind('<<ComboboxSelected>>', lambda e: self.load_region_stats())
        
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        columns = ('区域', '合同数', '合同总额', '开票总额', '回款总额')
        self.region_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.region_tree.heading(col, text=col)
            self.region_tree.column(col, width=150, anchor='center')
        
        self.region_tree.pack(fill='both', expand=True)
        
        chart_frame = ttk.LabelFrame(parent, text="区域统计图表")
        chart_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.region_fig = Figure(figsize=(12, 4), dpi=100)
        self.region_canvas = FigureCanvasTkAgg(self.region_fig, master=chart_frame)
        self.region_canvas.get_tk_widget().pack(fill='both', expand=True)
    
    def create_salesperson_tab(self, parent):
        filter_frame = ttk.Frame(parent)
        filter_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(filter_frame, text="年度:").pack(side='left', padx=2)
        self.sales_year_filter = ttk.Combobox(filter_frame, width=10, state='readonly')
        self.sales_year_filter['values'] = ['全部'] + [str(y) for y in range(2030, 2019, -1)]
        self.sales_year_filter.set('全部')
        self.sales_year_filter.pack(side='left', padx=2)
        self.sales_year_filter.bind('<<ComboboxSelected>>', lambda e: self.load_salesperson_stats())
        
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        columns = ('销售负责人', '合同数', '合同总额', '开票总额', '回款总额', '应收账款')
        self.sales_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.sales_tree.heading(col, text=col)
            self.sales_tree.column(col, width=150, anchor='center')
        
        self.sales_tree.pack(fill='both', expand=True)
        
        chart_frame = ttk.LabelFrame(parent, text="销售人员业绩图表")
        chart_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.sales_fig = Figure(figsize=(12, 4), dpi=100)
        self.sales_canvas = FigureCanvasTkAgg(self.sales_fig, master=chart_frame)
        self.sales_canvas.get_tk_widget().pack(fill='both', expand=True)
    
    # 合同管理相关方法
    def sort_contract_by_column(self, col):
        """合同列表排序"""
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        self.load_data()
    
    def clear_contract_search(self):
        """清除合同搜索"""
        self.contract_search_entry.delete(0, tk.END)
        self.contract_year_filter.set('全部')
        self.contract_region_filter.set('全部')
        self.contract_salesperson_filter.set('全部')
        self.load_data()
    
    def load_data(self):
        """加载合同数据"""
        year = None if self.contract_year_filter.get() == '全部' else self.contract_year_filter.get()
        region = None if self.contract_region_filter.get() == '全部' else self.contract_region_filter.get()
        salesperson = None if self.contract_salesperson_filter.get() == '全部' else self.contract_salesperson_filter.get()
        search_text = self.contract_search_entry.get().strip() or None
        
        rows = self.db.get_contracts(year, region, salesperson, search_text)
        
        # 清空列表
        for item in self.contract_tree.get_children():
            self.contract_tree.delete(item)
        
        # 插入数据
        for row in rows:
            values = []
            for i in range(29):  # 29个字段
                if i < len(row):
                    val = row[i]
                    # 格式化金额字段
                    if i in [12, 13, 22, 23, 24, 25, 26, 28]:  # 金额字段
                        if val is not None:
                            values.append(f'{val:,.2f}')
                        else:
                            values.append('')
                    else:
                        values.append(str(val) if val is not None else '')
                else:
                    values.append('')
            
            self.contract_tree.insert('', 'end', values=values)
        
        # 更新销售负责人筛选器
        self.update_salesperson_filter()
        
        # 加载统计数据
        self.load_yearly_stats()
        self.load_region_stats()
        self.load_salesperson_stats()
    
    def update_salesperson_filter(self):
        """更新销售负责人筛选器"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT 销售负责人 FROM contracts WHERE 销售负责人 IS NOT NULL AND 销售负责人 != ""')
        salespersons = ['全部'] + [row[0] for row in cursor.fetchall()]
        conn.close()
        self.contract_salesperson_filter['values'] = salespersons
    
    def add_contract(self):
        """添加合同"""
        self.show_contract_dialog()
    
    def edit_contract(self):
        """修改合同"""
        selected = self.contract_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择一个合同")
            return
        
        contract_no = self.contract_tree.item(selected[0])['values'][2]  # 合同编号在第3列
        self.show_contract_dialog(contract_no)
    
    def delete_contract(self):
        """删除合同"""
        selected = self.contract_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择一个合同")
            return
        
        contract_no = self.contract_tree.item(selected[0])['values'][2]
        contract_name = self.contract_tree.item(selected[0])['values'][8]
        
        if not messagebox.askyesno("确认删除", f"确定要删除合同 '{contract_no}' 吗？\n\n合同名称: {contract_name}"):
            return
        
        if self.db.delete_contract(contract_no):
            messagebox.showinfo("成功", "合同删除成功")
            self.load_data()
        else:
            messagebox.showerror("错误", "合同删除失败")
    
    def show_contract_dialog(self, contract_no=None):
        """显示合同对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("修改合同" if contract_no else "添加合同")
        dialog.geometry("800x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 创建画布和滚动条
        canvas = tk.Canvas(dialog)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 字段定义
        fields = {}
        field_defs = [
            ('序号', '序号', 'entry'),
            ('下单日期', '下单日期', 'date'),
            ('合同编号', '合同编号', 'entry'),
            ('项目代码', '项目代码', 'entry'),
            ('是否变更', '是否变更', 'entry'),
            ('合同评审日期', '合同评审日期', 'date'),
            ('合同签字日期', '合同签字日期', 'date'),
            ('crm日期', 'crm日期', 'date'),
            ('合同名称', '合同名称', 'entry'),
            ('对方单位名称', '对方单位名称', 'entry'),
            ('区域', '区域', 'combo'),
            ('销售负责人', '销售负责人', 'entry'),
            ('参考金额', '参考金额', 'entry'),
            ('合同额', '合同额', 'entry'),
            ('联系人', '联系人', 'entry'),
            ('联系电话', '联系电话', 'entry'),
            ('合同内容', '合同内容', 'entry'),
            ('到款情况', '到款情况', 'entry'),
            ('合同起始日期', '合同起始日期', 'date'),
            ('合同终止日期', '合同终止日期', 'date'),
            ('开票日期', '开票日期', 'date'),
            ('开票金额', '开票金额', 'entry'),
            ('开票余额', '开票余额', 'entry'),
            ('到款金额', '到款金额', 'entry'),
            ('合同余额', '合同余额', 'entry'),
            ('应收账款', '应收账款', 'entry'),
            ('备注', '备注', 'entry'),
            ('项目预算', '项目预算', 'entry'),
            ('设备数量', '设备数量', 'entry'),
        ]
        
        row = 0
        for label_text, field_name, field_type in field_defs:
            ttk.Label(scrollable_frame, text=f"{label_text}:").grid(row=row, column=0, sticky='e', pady=3, padx=5)
            
            if field_type == 'date':
                fields[field_name] = DateEntry(scrollable_frame, width=27, date_pattern='yyyy-mm-dd', locale='zh_CN')
            elif field_type == 'combo':
                fields[field_name] = ttk.Combobox(scrollable_frame, width=28, values=self.regions, state='readonly')
            else:
                fields[field_name] = ttk.Entry(scrollable_frame, width=30)
            
            fields[field_name].grid(row=row, column=1, pady=3, padx=5, sticky='w')
            row += 1
        
        # 如果是修改，填充数据
        if contract_no:
            contract = self.db.get_contract_by_no(contract_no)
            if contract:
                for i, (label_text, field_name, field_type) in enumerate(field_defs):
                    if i < len(contract):
                        val = contract[i]
                        if val is not None:
                            if field_type == 'date' and val:
                                try:
                                    fields[field_name].set_date(val)
                                except:
                                    pass
                            elif field_type == 'combo':
                                fields[field_name].set(val)
                            else:
                                fields[field_name].insert(0, str(val))
                
                fields['合同编号'].config(state='disabled')
        
        # 按钮
        btn_frame = ttk.Frame(scrollable_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        def save():
            data = {}
            for label_text, field_name, field_type in field_defs:
                val = fields[field_name].get()
                if field_type == 'entry' and field_name in ['序号', '参考金额', '合同额', '开票金额', '开票余额', '到款金额', '合同余额', '应收账款', '项目预算', '设备数量']:
                    try:
                        data[field_name] = float(val) if val and field_name != '序号' and field_name != '设备数量' else (int(val) if val else None)
                    except:
                        data[field_name] = None
                else:
                    data[field_name] = val if val else None
            
            if not data.get('合同编号'):
                messagebox.showwarning("警告", "合同编号不能为空")
                return
            
            if contract_no:
                if self.db.update_contract(contract_no, data):
                    messagebox.showinfo("成功", "合同修改成功")
                    dialog.destroy()
                    self.load_data()
                else:
                    messagebox.showerror("错误", "合同修改失败")
            else:
                success, msg = self.db.add_contract(data)
                if success:
                    messagebox.showinfo("成功", "合同添加成功")
                    dialog.destroy()
                    self.load_data()
                else:
                    messagebox.showwarning("警告", msg)
        
        ttk.Button(btn_frame, text="保存", command=save).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side='left', padx=10)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def import_contracts(self):
        """导入合同数据"""
        file_path = filedialog.askopenfilename(
            title="选择合同数据文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            count, dup_count, error_count = self.db.import_contracts_from_file(file_path)
            msg = f"导入完成！\n\n成功: {count} 条\n重复跳过: {dup_count} 条\n失败: {error_count} 条"
            messagebox.showinfo("导入结果", msg)
            self.load_data()
        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}")
    
    # 发票管理相关方法
    def sort_invoice_by_column(self, col):
        """发票列表排序"""
        # 简单实现：重新加载数据
        self.load_invoices()
    
    def clear_invoice_search(self):
        """清除发票搜索"""
        self.invoice_search_entry.delete(0, tk.END)
        self.load_invoices()
    
    def load_invoices(self):
        """加载发票数据"""
        search_text = self.invoice_search_entry.get().strip() or None
        rows = self.db.get_invoices_new(search_text)
        
        for item in self.invoice_tree.get_children():
            self.invoice_tree.delete(item)
        
        for row in rows:
            # row结构: id, 开票日期, 合同号, 付款单位名称, 代码, 发票金额, 发票项目, 类型, 发票类型, 除税, 备注, 数据哈希, 创建时间
            values = []
            for i in range(1, 11):  # 跳过id和最后的字段
                if i < len(row):
                    val = row[i]
                    if i == 5:  # 发票金额
                        values.append(f'{val:,.2f}' if val else '0.00')
                    elif i == 9:  # 除税
                        values.append(f'{val:,.2f}' if val else '')
                    else:
                        values.append(str(val) if val else '')
                else:
                    values.append('')
            
            self.invoice_tree.insert('', 'end', values=values)
    
    def add_invoice(self):
        """添加发票"""
        self.show_invoice_dialog()
    
    def edit_invoice(self):
        """修改发票"""
        selected = self.invoice_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择一条发票记录")
            return
        
        # 获取发票ID（需要从数据库查询）
        values = self.invoice_tree.item(selected[0])['values']
        # 这里简化处理，实际应该根据唯一标识查询
        messagebox.showinfo("提示", "请双击发票记录进行修改")
    
    def delete_invoice(self):
        """删除发票"""
        selected = self.invoice_tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择一条发票记录")
            return
        
        if not messagebox.askyesno("确认删除", "确定要删除选中的发票记录吗？"):
            return
        
        # 简化处理
        messagebox.showinfo("提示", "删除功能需要发票ID，请双击记录进行操作")
    
    def show_invoice_dialog(self, invoice_id=None):
        """显示发票对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("修改发票" if invoice_id else "添加发票")
        dialog.geometry("500x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill='both', expand=True)
        
        fields = {}
        field_defs = [
            ('开票日期', '开票日期', 'date'),
            ('合同号', '合同号', 'entry'),
            ('付款单位名称', '付款单位名称', 'entry'),
            ('代码', '代码', 'entry'),
            ('发票金额', '发票金额', 'entry'),
            ('发票项目', '发票项目', 'entry'),
            ('类型', '类型', 'entry'),
            ('发票类型', '发票类型', 'entry'),
            ('除税', '除税', 'entry'),
            ('备注', '备注', 'entry'),
        ]
        
        row = 0
        for label_text, field_name, field_type in field_defs:
            ttk.Label(frame, text=f"{label_text}:").grid(row=row, column=0, sticky='e', pady=5)
            
            if field_type == 'date':
                fields[field_name] = DateEntry(frame, width=27, date_pattern='yyyy-mm-dd', locale='zh_CN')
            else:
                fields[field_name] = ttk.Entry(frame, width=30)
            
            fields[field_name].grid(row=row, column=1, pady=5)
            row += 1
        
        # 如果是修改，填充数据
        if invoice_id:
            invoice = self.db.get_invoice_by_id(invoice_id)
            if invoice:
                for i, (label_text, field_name, field_type) in enumerate(field_defs):
                    if i + 1 < len(invoice):
                        val = invoice[i + 1]
                        if val:
                            if field_type == 'date':
                                try:
                                    fields[field_name].set_date(val)
                                except:
                                    pass
                            else:
                                fields[field_name].insert(0, str(val))
        
        # 按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)
        
        def save():
            data = {}
            for label_text, field_name, field_type in field_defs:
                val = fields[field_name].get()
                if field_type == 'entry' and field_name in ['发票金额', '除税']:
                    try:
                        data[field_name] = float(val) if val else None
                    except:
                        data[field_name] = None
                else:
                    data[field_name] = val if val else None
            
            if invoice_id:
                if self.db.update_invoice_new(invoice_id, data):
                    messagebox.showinfo("成功", "发票修改成功")
                    dialog.destroy()
                    self.load_invoices()
                else:
                    messagebox.showerror("错误", "发票修改失败")
            else:
                success, msg = self.db.add_invoice_new(data)
                if success:
                    messagebox.showinfo("成功", "发票添加成功")
                    dialog.destroy()
                    self.load_invoices()
                else:
                    messagebox.showwarning("警告", msg)
        
        ttk.Button(btn_frame, text="保存", command=save).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side='left', padx=10)
    
    def import_invoices(self):
        """导入发票数据"""
        file_path = filedialog.askopenfilename(
            title="选择发票数据文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            count, dup_count, error_count = self.db.import_invoices_from_file(file_path)
            msg = f"导入完成！\n\n成功: {count} 条\n重复跳过: {dup_count} 条\n失败: {error_count} 条"
            messagebox.showinfo("导入结果", msg)
            self.load_invoices()
        except Exception as e:
            messagebox.showerror("错误", f"导入失败: {str(e)}")
    
    # 其他方法（预警、统计等）
    def load_warning_data(self):
        """加载到期预警数据"""
        for item in self.warning_tree.get_children():
            self.warning_tree.delete(item)
        
        today = datetime.now()
        
        expiring = self.db.get_expiring_contracts(30)
        expired = self.db.get_expired_contracts()
        
        all_contracts = []
        
        for row in expiring:
            end_date = row[20] if len(row) > 20 else ''  # 合同终止日期
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                days_left = (end_dt - today).days
            except:
                days_left = 0
            all_contracts.append((row, days_left, 'expiring'))
        
        for row in expired:
            end_date = row[20] if len(row) > 20 else ''
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                days_left = (end_dt - today).days
            except:
                days_left = -999
            all_contracts.append((row, days_left, 'expired'))
        
        all_contracts.sort(key=lambda x: x[1])
        
        # 获取过滤选项
        filter_type = self.warning_filter_var.get() if hasattr(self, 'warning_filter_var') else 'all'
        
        # 获取搜索文本
        search_text = self.warning_search_entry.get().strip() if hasattr(self, 'warning_search_entry') else ''
        
        for row, days_left, status in all_contracts:
            合同编号 = row[3] if len(row) > 3 else ''
            合同名称 = row[9] if len(row) > 9 else ''
            对方单位 = row[10] if len(row) > 10 else ''
            区域 = row[11] if len(row) > 11 else ''
            销售负责人 = row[12] if len(row) > 12 else ''
            合同额 = row[14] if len(row) > 14 else 0
            终止日期 = row[20] if len(row) > 20 else ''
            开票金额 = row[23] if len(row) > 23 else 0
            到款金额 = row[25] if len(row) > 25 else 0
            应收账款 = row[27] if len(row) > 27 else 0
            
            # 确定预警等级
            if days_left >= 0:
                tag = 'warning_yellow'
            elif days_left >= -30:
                tag = 'warning_orange'
            else:
                tag = 'warning_red'
            
            # 快速过滤
            if filter_type == 'yellow' and tag != 'warning_yellow':
                continue
            elif filter_type == 'orange' and tag != 'warning_orange':
                continue
            elif filter_type == 'red' and tag != 'warning_red':
                continue
            
            # 搜索过滤
            if search_text:
                search_lower = search_text.lower()
                if not (search_lower in 合同编号.lower() or 
                       search_lower in 合同名称.lower() or 
                       search_lower in 对方单位.lower() or 
                       search_lower in 区域.lower() or 
                       search_lower in (销售负责人 or '').lower()):
                    continue
            
            self.warning_tree.insert('', 'end', values=(
                合同编号, 合同名称 or '', 对方单位, 区域, 销售负责人 or '',
                f'{合同额:,.2f}' if 合同额 else '0.00',
                终止日期 or '', days_left,
                f'{开票金额:,.2f}' if 开票金额 else '0.00',
                f'{到款金额:,.2f}' if 到款金额 else '0.00',
                f'{应收账款:,.2f}' if 应收账款 else '0.00'
            ), tags=(tag,))
    
    def clear_warning_search(self):
        """清除预警搜索"""
        if hasattr(self, 'warning_search_entry'):
            self.warning_search_entry.delete(0, tk.END)
        if hasattr(self, 'warning_filter_var'):
            self.warning_filter_var.set('all')
        self.load_warning_data()
    
    def check_expiring_contracts(self):
        """检查即将到期的合同"""
        expiring = self.db.get_expiring_contracts(30)
        
        if expiring:
            count = len(expiring)
            message = f"当前有 {count} 份合同将在30天内到期！\n\n"
            message += "即将到期的合同：\n"
            
            for i, row in enumerate(expiring[:5]):
                合同编号 = row[3] if len(row) > 3 else ''
                合同名称 = row[9] if len(row) > 9 else ''
                终止日期 = row[20] if len(row) > 20 else ''
                message += f"{i+1}. {合同编号} - {合同名称} (到期: {终止日期})\n"
            
            if count > 5:
                message += f"\n... 还有 {count - 5} 份合同即将到期"
            
            messagebox.showwarning("到期预警", message)
    
    def load_yearly_stats(self):
        """加载年度统计"""
        rows = self.db.get_yearly_stats()
        
        for item in self.yearly_tree.get_children():
            self.yearly_tree.delete(item)
        
        for row in rows:
            year, count, total, invoice, payment = row
            receivable = (invoice or 0) - (payment or 0)
            self.yearly_tree.insert('', 'end', values=(
                year, count,
                f'{total:,.2f}' if total else '0.00',
                f'{invoice:,.2f}' if invoice else '0.00',
                f'{payment:,.2f}' if payment else '0.00',
                f'{receivable:,.2f}'
            ))
        
        self.draw_yearly_chart(rows)
    
    def draw_yearly_chart(self, rows):
        """绘制年度统计图表"""
        self.yearly_fig.clear()
        
        if not rows:
            self.yearly_canvas.draw()
            return
        
        years = [row[0] for row in rows][::-1]
        invoices = [row[3] or 0 for row in rows][::-1]
        payments = [row[4] or 0 for row in rows][::-1]
        
        ax = self.yearly_fig.add_subplot(111)
        
        x = range(len(years))
        width = 0.35
        
        bars1 = ax.bar([i - width/2 for i in x], invoices, width, label='开票总额', color='#3498db')
        bars2 = ax.bar([i + width/2 for i in x], payments, width, label='回款总额', color='#2ecc71')
        
        ax.set_xlabel('年度')
        ax.set_ylabel('金额（元）')
        ax.set_title('年度开票与回款统计')
        ax.set_xticks(x)
        ax.set_xticklabels(years)
        ax.legend()
        
        self.yearly_fig.tight_layout()
        self.yearly_canvas.draw()
    
    def load_region_stats(self):
        """加载区域统计"""
        year = None if self.region_year_filter.get() == '全部' else self.region_year_filter.get()
        rows = self.db.get_region_stats(year)
        
        for item in self.region_tree.get_children():
            self.region_tree.delete(item)
        
        for row in rows:
            region, count, total, invoice, payment = row
            self.region_tree.insert('', 'end', values=(
                region, count,
                f'{total:,.2f}' if total else '0.00',
                f'{invoice:,.2f}' if invoice else '0.00',
                f'{payment:,.2f}' if payment else '0.00'
            ))
        
        self.draw_region_chart(rows)
    
    def draw_region_chart(self, rows):
        """绘制区域统计图表"""
        self.region_fig.clear()
        
        if not rows:
            self.region_canvas.draw()
            return
        
        regions = [row[0] for row in rows]
        totals = [row[2] or 0 for row in rows]
        
        ax = self.region_fig.add_subplot(111)
        
        x = range(len(regions))
        bars = ax.bar(x, totals, color='#e74c3c')
        
        ax.set_xlabel('区域')
        ax.set_ylabel('合同总额（元）')
        ax.set_title('区域合同统计')
        ax.set_xticks(x)
        ax.set_xticklabels(regions, rotation=45, ha='right')
        
        self.region_fig.tight_layout()
        self.region_canvas.draw()
    
    def load_salesperson_stats(self):
        """加载销售人员统计"""
        year = None if self.sales_year_filter.get() == '全部' else self.sales_year_filter.get()
        rows = self.db.get_salesperson_stats(year)
        
        for item in self.sales_tree.get_children():
            self.sales_tree.delete(item)
        
        for row in rows:
            salesperson, count, total, invoice, payment = row
            receivable = (invoice or 0) - (payment or 0)
            self.sales_tree.insert('', 'end', values=(
                salesperson or '未指定', count,
                f'{total:,.2f}' if total else '0.00',
                f'{invoice:,.2f}' if invoice else '0.00',
                f'{payment:,.2f}' if payment else '0.00',
                f'{receivable:,.2f}'
            ))
        
        self.draw_salesperson_chart(rows)
    
    def draw_salesperson_chart(self, rows):
        """绘制销售人员统计图表"""
        self.sales_fig.clear()
        
        if not rows:
            self.sales_canvas.draw()
            return
        
        salespersons = [row[0] or '未指定' for row in rows]
        totals = [row[2] or 0 for row in rows]
        
        ax = self.sales_fig.add_subplot(111)
        
        x = range(len(salespersons))
        bars = ax.bar(x, totals, color='#9b59b6')
        
        ax.set_xlabel('销售负责人')
        ax.set_ylabel('合同总额（元）')
        ax.set_title('销售人员业绩统计')
        ax.set_xticks(x)
        ax.set_xticklabels(salespersons, rotation=45, ha='right')
        
        self.sales_fig.tight_layout()
        self.sales_canvas.draw()
    
    # 数据导出和备份
    def backup_database(self):
        """备份数据库"""
        file_path = filedialog.asksaveasfilename(
            title="保存备份文件",
            defaultextension=".db",
            filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import shutil
            shutil.copy2(DB_PATH, file_path)
            messagebox.showinfo("成功", f"数据库已备份到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"备份失败: {str(e)}")
    
    def restore_database(self):
        """恢复数据库"""
        if not messagebox.askyesno("确认", "恢复数据库将覆盖当前数据，是否继续？"):
            return
        
        file_path = filedialog.askopenfilename(
            title="选择备份文件",
            filetypes=[("数据库文件", "*.db"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import shutil
            shutil.copy2(file_path, DB_PATH)
            messagebox.showinfo("成功", "数据库已恢复，程序将重新启动")
            self.root.destroy()
        except Exception as e:
            messagebox.showerror("错误", f"恢复失败: {str(e)}")
    
    def export_contracts(self):
        """导出合同"""
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "合同列表"
            
            headers = ('序号', '下单日期', '合同编号', '项目代码', '是否变更', '合同评审日期',
                      '合同签字日期', 'crm日期', '合同名称', '对方单位名称', '区域', '销售负责人',
                      '参考金额', '合同额', '联系人', '联系电话', '合同内容', '到款情况',
                      '合同起始日期', '合同终止日期', '开票日期', '开票金额', '开票余额',
                      '到款金额', '合同余额', '应收账款', '备注', '项目预算', '设备数量')
            ws.append(headers)
            
            for item in self.contract_tree.get_children():
                values = self.contract_tree.item(item)['values']
                ws.append(values)
            
            wb.save(file_path)
            messagebox.showinfo("成功", f"合同数据已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def export_invoices(self):
        """导出发票"""
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "发票列表"
            
            headers = ('开票日期', '合同号', '付款单位名称', '代码', '发票金额', '发票项目',
                      '类型', '发票类型', '除税', '备注')
            ws.append(headers)
            
            for item in self.invoice_tree.get_children():
                values = self.invoice_tree.item(item)['values']
                ws.append(values)
            
            wb.save(file_path)
            messagebox.showinfo("成功", f"发票数据已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def export_warning_list(self):
        """导出预警列表"""
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "到期预警"
            
            headers = ['合同编号', '合同名称', '对方单位', '区域', '销售负责人',
                      '合同额', '终止日期', '剩余天数', '开票金额', '到款金额', '应收账款']
            ws.append(headers)
            
            for item in self.warning_tree.get_children():
                values = self.warning_tree.item(item)['values']
                ws.append(values)
            
            wb.save(file_path)
            messagebox.showinfo("成功", f"预警列表已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def export_receivable_excel(self, year):
        """导出应收账款"""
        file_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{year}年应收账款"
            
            headers = ['合同编号', '合同名称', '对方单位', '区域', '销售负责人', '合同额',
                      '签字日期', '账龄(天)', '开票金额', '到款金额', '应收金额', '催款状态']
            ws.append(headers)
            
            if year in self.receivable_trees:
                tree = self.receivable_trees[year]
                for item in tree.get_children():
                    values = tree.item(item)['values'][:12]
                    ws.append(values)
            
            wb.save(file_path)
            messagebox.showinfo("成功", f"应收账款已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def show_about(self):
        """显示关于"""
        messagebox.showinfo("关于", 
            "合同管理系统 v5.0\n\n"
            "功能：\n"
            "- 合同管理（29字段）\n"
            "- 发票管理\n"
            "- 搜索和排序\n"
            "- 到期合同预警\n"
            "- 应收账款催款\n"
            "- 统计分析\n\n"
            "数据存储位置：\n"
            f"{DATA_DIR}")


def main():
    root = tk.Tk()
    app = ContractManagerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
